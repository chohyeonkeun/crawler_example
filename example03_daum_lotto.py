import requests

from bs4 import BeautifulSoup

from openpyxl import Workbook

# 1. 로또 번호 크롤러
# 1) 1회차분 크롤링
# 2) 다회차 크롤링

# 2. 엑셀에 데이터 저장
# 1) 아무값이나 엑셀에 ㅓㅅ보기
# 2) 특정 회차 번호 저장해보기
# 3) 다회차 저장해보기

# 1 ~ 855회차까지 로또 번호 수집
# 엑셀 파일 저장
# 1 : 23, 2 : 10, 3 : 7, 4 : 12

wb = Workbook()
worksheet = wb.active
worksheet.title = 'Count'

worksheet2 = wb.create_sheet(title = "List")

numbers_count = {x : 0 for x in range(1, 46)}   # numbers_count = {1 : 0, 2 : 0, 3 : 0, ... 45 : 0}

for num in range(1, 5):
    url = "https://search.daum.net/search?w=tot&rtmaxcoll=LOT&DA=LOT&q=" + str(num) + "%ED%9A%8C%EC%B0%A8%20%EB%A1%9C%EB%98%90"

    req = requests.get(url)

    if req.status_code == requests.codes.ok:
        print(f'{num}회차 접속 성공')

        html = BeautifulSoup(req.text, 'html.parser')
        numbers = html.select('.lottonum .img_lotto')
        print(numbers)
        worksheet2.cell(row = num, column = 1,  value = f'{num}회차')
        for index, number in enumerate(numbers):
            try:
                # 당첨 번호 카운팅
                numbers_count[int(number.text)] += 1

                worksheet2.cell(row = num, column = index + 2, value = int(number.text))

            except:
                pass

for number in numbers_count:
    print(number, ":", numbers_count[number])

for row in range(1, 46):
    worksheet.cell(row = row, column = 1, value = row)
    worksheet.cell(row = row, column = 2, value = numbers_count[row])
wb.save(filename = 'lotto.xlsx')




'''
1회차  10 23 ...
2회차
3회차

'''
