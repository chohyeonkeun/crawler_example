# pip install openpyxl

from openpyxl import Workbook

# Workbook : 파일
# Worksheet : 탭
# 내용을 저장하기 위해 메모리 상에 파일 만들기
wb = Workbook()

ws1 = wb.active

ws1.cell(row = 1, column = 1, value = 'test') # == ws1['A1'] = 'test'
ws2 = wb.create_sheet(title = 'second sheet')
ws2.cell(row = 3, column = 7, value = '두번째 시트')
wb.save('test.xlsx')

