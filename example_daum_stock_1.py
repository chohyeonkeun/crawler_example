"""
메뉴 1 : 조회 급등 항목 10개의 현재가 출력

메뉴 2 : 1) 종목 코드를 입력받고 2) 해당 종목의 1페이지 주가 출력

메뉴 2-2 : 데이터 엑셀로 저장하기

1. 메뉴 구현하기
2. 1번 메뉴에 대한 크롤러
    2-1. 조회 급등 항목 10개 찾기
    2-2. 그 항목 1개에 대한 데이터 찾아서 출력
    2-3. 총 10개 항목에 대한 데이터 찾아서 출력
3. 2번 메뉴에 대한 크롤러
    3-1. 아무 항목이나 고정 - 1페이지 주가 출력
    3-2. 사용자에게 종목 코드 입력받아서 출력하기
"""

import json
import requests
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "rank_list"

def get_stock_price(stock_code):
    url = "http://finance.daum.net/api/quote/"+stock_code+"/days?symbolCode="+stock_code+"&page=2&perPage=10&pagination=true"
    custom_header = {
        'referer': "http://finance.daum.net/domestic",
        'user-agent': "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.75 Safari/537.36",
    }
    req = requests.get(url, headers = custom_header)
    data = json.loads(req.text)
    current_data = data['data'][0]['tradePrice']  # data = {data : [{symbolCode: "", tradePrice:"12650"} {} [} [}]}
    return current_data

url = "http://finance.daum.net/api/search/ranks?limit=10"
custom_header = {
    'referer' : "http://finance.daum.net/domestic",
    'user-agent' : "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.75 Safari/537.36",
}

req = requests.get(url, headers = custom_header)

if req.status_code == requests.codes.ok:
    print('접속 성공')
    # print(req.text)
    rank_data = json.loads(req.text)
    # print(rank_data)
    for idx, rank in enumerate(rank_data['data'], 1):
        current_price = get_stock_price(rank['symbolCode'])

        print(rank['rank'], rank['name'], rank['symbolCode'], rank['tradePrice'], current_price)

        ws.cell(row=idx, column=1, value=idx)
        ws.cell(row=idx, column=2, value=rank['name'])
        ws.cell(row=idx, column=3, value=rank['symbolCode'])
        ws.cell(row=idx, column=4, value=rank['tradePrice'])
        ws.cell(row=idx, column=5, value=current_price)

    wb.save(filename="daum_stock_rank.xlsx")
